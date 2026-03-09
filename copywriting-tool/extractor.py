#!/usr/bin/env python3
"""
Next.js Copywriting Extractor Tool
A reusable tool for extracting copywriting from Next.js i18n projects
Supports next-intl, next-i18next, and custom i18n implementations
"""

import json
import csv
import yaml
import argparse
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime
import re

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class CopywritingExtractor:
    """Extract copywriting from Next.js i18n JSON files"""
    
    def __init__(self, config_path: Optional[str] = None):
        self.config = self._load_config(config_path)
        self.copywriting_rows: List[Dict[str, str]] = []
        self.locale_data: Dict[str, Any] = {}
        
    def _load_config(self, config_path: Optional[str]) -> dict:
        """Load configuration from YAML file"""
        default_config = {
            'input': {
                'locale_dir': 'messages',
                'locales': ['en'],
                'file_pattern': '{locale}.json',
            },
            'output': {
                'filename': 'Copywriting_{date}',
                'formats': ['csv', 'xlsx'],
                'columns': ['route', 'page', 'section', 'content_type', 'key'],
            },
            'extraction': {
                'route_mapping': {},
                'page_types': {
                    'b2c': [],
                    'b2b': [],
                    'global': ['Navigation', 'Footer', 'StickyBanner', 'PopUp']
                },
                'skip_keys': [],
                'flatten_nested': True,
                'max_depth': 10,
            },
            'styling': {
                'header_color': '193CB8',
                'header_text_color': 'FFFFFF',
                'alternate_row_color': 'F5F5F5',
            }
        }
        
        if config_path and Path(config_path).exists():
            with open(config_path, 'r', encoding='utf-8') as f:
                user_config = yaml.safe_load(f) or {}
                return self._merge_config(default_config, user_config)
        
        return default_config
    
    def _merge_config(self, default: dict, user: dict) -> dict:
        """Recursively merge user config with defaults"""
        result = default.copy()
        for key, value in user.items():
            if key in result and isinstance(result[key], dict) and isinstance(value, dict):
                result[key] = self._merge_config(result[key], value)
            else:
                result[key] = value
        return result
    
    def load_locales(self, workspace_root: Path) -> None:
        """Load locale JSON files"""
        locale_dir = workspace_root / self.config['input']['locale_dir']
        file_pattern = self.config['input']['file_pattern']
        
        for locale in self.config['input']['locales']:
            locale_file = locale_dir / file_pattern.format(locale=locale)
            if locale_file.exists():
                with open(locale_file, 'r', encoding='utf-8') as f:
                    self.locale_data[locale] = json.load(f)
                print(f"✓ Loaded {locale} locale from {locale_file}")
            else:
                print(f"⚠ Warning: {locale_file} not found")
    
    def extract_all(self, locale: str = 'en') -> List[Dict[str, str]]:
        """Extract all copywriting from specified locale"""
        if locale not in self.locale_data:
            raise ValueError(f"Locale '{locale}' not loaded")
        
        data = self.locale_data[locale]
        self.copywriting_rows = []
        
        # Extract based on configuration
        for section_key, section_data in data.items():
            route = self._determine_route(section_key)
            page_type = self._determine_page_type(section_key)
            
            if section_key in self.config['extraction']['skip_keys']:
                continue
            
            self._extract_section(
                section_data=section_data,
                route=route,
                page=page_type,
                section=section_key,
                prefix='',
                depth=0
            )
        
        return self.copywriting_rows
    
    def _determine_route(self, section_key: str) -> str:
        """Determine route from section key using mapping"""
        route_mapping = self.config['extraction']['route_mapping']
        
        # Check explicit mapping
        if section_key in route_mapping:
            return route_mapping[section_key]
        
        # Check for common patterns
        if section_key in self.config['extraction']['page_types']['global']:
            return 'Global'
        
        if 'B2B' in section_key or 'Enterprise' in section_key or section_key == 'TechStack':
            return '/enterprise-solutions'
        
        if 'Terms' in section_key or 'Privacy' in section_key or 'Cancellation' in section_key:
            # Convert section key to route
            route = re.sub(r'([A-Z])', r'-\1', section_key).lower().strip('-')
            return f'/{route}'
        
        return '/'
    
    def _determine_page_type(self, section_key: str) -> str:
        """Determine page type (b2c, b2b, or N/A)"""
        page_types = self.config['extraction']['page_types']
        
        for page_type, sections in page_types.items():
            if section_key in sections:
                return 'N/A' if page_type == 'global' else page_type
        
        # Infer from section name
        if any(keyword in section_key for keyword in ['B2B', 'Enterprise', 'TechStack', 'Infrastructure', 'Partnership']):
            return 'b2b'
        
        return 'b2c'
    
    def _extract_section(
        self,
        section_data: Any,
        route: str,
        page: str,
        section: str,
        prefix: str,
        depth: int
    ) -> None:
        """Recursively extract copywriting from a section"""
        max_depth = self.config['extraction']['max_depth']
        
        if depth > max_depth:
            return
        
        if isinstance(section_data, dict):
            for key, value in section_data.items():
                new_prefix = f"{prefix}_{key}" if prefix else key
                
                if isinstance(value, (dict, list)):
                    if self.config['extraction']['flatten_nested']:
                        self._extract_section(value, route, page, section, new_prefix, depth + 1)
                    else:
                        # Add the entire nested object as JSON string
                        self.add_row(route, page, section, prefix or 'nested', json.dumps(value, ensure_ascii=False))
                else:
                    # Leaf node - add to rows
                    content_type = self._infer_content_type(new_prefix, key)
                    self.add_row(route, page, section, content_type, str(value))
        
        elif isinstance(section_data, list):
            # Handle arrays
            for i, item in enumerate(section_data):
                new_prefix = f"{prefix}_{i+1}" if prefix else f"item_{i+1}"
                if isinstance(item, (dict, list)):
                    self._extract_section(item, route, page, section, new_prefix, depth + 1)
                else:
                    content_type = self._infer_content_type(new_prefix, str(i))
                    self.add_row(route, page, section, content_type, str(item))
        
        else:
            # Primitive value
            content_type = self._infer_content_type(prefix, section)
            self.add_row(route, page, section, content_type, str(section_data))
    
    def _infer_content_type(self, key_path: str, last_key: str) -> str:
        """Infer content type from key name"""
        key_lower = key_path.lower()
        
        # Button patterns
        if any(keyword in key_lower for keyword in ['button', 'btn', 'cta']):
            return 'button'
        
        # Heading patterns
        if any(keyword in key_lower for keyword in ['heading', 'title', 'headline']):
            return 'heading'
        
        # Subheading patterns
        if any(keyword in key_lower for keyword in ['subheading', 'subtitle', 'subhead']):
            return 'subheading'
        
        # Description patterns
        if any(keyword in key_lower for keyword in ['description', 'desc', 'content', 'text', 'body']):
            return 'description'
        
        # Badge/Label patterns
        if any(keyword in key_lower for keyword in ['badge', 'label', 'tag']):
            return 'badge'
        
        # Question/Answer patterns
        if 'question' in key_lower:
            return 'faq_question'
        if 'answer' in key_lower:
            return 'faq_answer'
        
        # Link patterns
        if any(keyword in key_lower for keyword in ['link', 'url', 'href']):
            return 'link'
        
        # Placeholder patterns
        if 'placeholder' in key_lower:
            return 'placeholder'
        
        # Highlighted word
        if 'highlighted' in key_lower or 'highlight' in key_lower:
            return 'highlighted_word'
        
        # Default
        return last_key.replace('_', ' ').title()
    
    def add_row(self, route: str, page: str, section: str, content_type: str, text: str) -> None:
        """Add a copywriting row"""
        if text and text.strip() and text != 'null':
            row = {
                'route': route,
                'page': page,
                'section': section,
                'content_type': content_type,
                'key': text.strip()
            }
            
            # Add locale column if extracting multiple languages
            if len(self.config['input']['locales']) > 1:
                row['locale'] = self.current_locale
            
            self.copywriting_rows.append(row)
    
    def generate_csv(self, output_path: Path) -> None:
        """Generate CSV file"""
        columns = self.config['output']['columns'].copy()
        if len(self.config['input']['locales']) > 1:
            columns.append('locale')
        
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=columns)
            writer.writeheader()
            writer.writerows(self.copywriting_rows)
        
        print(f"✓ CSV file created: {output_path}")
    
    def generate_xlsx(self, output_path: Path) -> None:
        """Generate styled XLSX file"""
        if not OPENPYXL_AVAILABLE:
            print("⚠ Skipping XLSX generation (openpyxl not installed)")
            print("  Install with: pip install openpyxl")
            return
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Copywriting"
        
        columns = self.config['output']['columns'].copy()
        if len(self.config['input']['locales']) > 1:
            columns.append('locale')
        
        # Styling
        header_color = self.config['styling']['header_color']
        header_text_color = self.config['styling']['header_text_color']
        
        # Add headers
        for col_num, header in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color=header_text_color)
            cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(
                bottom=Side(style='thin', color='000000')
            )
        
        # Add data with alternating row colors
        alternate_color = self.config['styling'].get('alternate_row_color', 'F5F5F5')
        
        for row_num, row_data in enumerate(self.copywriting_rows, 2):
            for col_num, col_name in enumerate(columns, 1):
                value = row_data.get(col_name, '')
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                
                # Alternate row coloring
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color=alternate_color, end_color=alternate_color, fill_type="solid")
        
        # Adjust column widths
        column_widths = {
            'route': 30,
            'page': 15,
            'section': 25,
            'content_type': 25,
            'key': 70,
            'locale': 10,
        }
        
        for col_num, col_name in enumerate(columns, 1):
            col_letter = openpyxl.utils.get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = column_widths.get(col_name, 20)
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Set row height for header
        ws.row_dimensions[1].height = 30
        
        wb.save(output_path)
        print(f"✓ XLSX file created: {output_path}")


def main():
    parser = argparse.ArgumentParser(
        description='Extract copywriting from Next.js i18n JSON files',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Extract with default config
  python extractor.py
  
  # Use custom config file
  python extractor.py --config my-config.yml
  
  # Specify workspace and output directory
  python extractor.py --workspace /path/to/project --output ./exports
  
  # Extract specific locale
  python extractor.py --locale en
  
  # Generate only CSV
  python extractor.py --format csv
        """
    )
    
    parser.add_argument(
        '--config', '-c',
        help='Path to configuration YAML file',
        default='copywriting.config.yml'
    )
    
    parser.add_argument(
        '--workspace', '-w',
        help='Workspace root directory (default: current directory)',
        default='.'
    )
    
    parser.add_argument(
        '--output', '-o',
        help='Output directory for generated files',
        default='.'
    )
    
    parser.add_argument(
        '--locale', '-l',
        help='Locale to extract (default: en)',
        default='en'
    )
    
    parser.add_argument(
        '--format', '-f',
        choices=['csv', 'xlsx', 'both'],
        help='Output format (default: both)',
        default='both'
    )
    
    parser.add_argument(
        '--prefix', '-p',
        help='Output filename prefix',
        default='Copywriting'
    )
    
    args = parser.parse_args()
    
    # Paths
    workspace_root = Path(args.workspace).resolve()
    output_dir = Path(args.output).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # Config path
    config_path = args.config
    if not Path(config_path).is_absolute():
        config_path = workspace_root / config_path
    
    print("🚀 Next.js Copywriting Extractor\n")
    print(f"Workspace: {workspace_root}")
    print(f"Output: {output_dir}\n")
    
    # Extract
    extractor = CopywritingExtractor(str(config_path) if config_path.exists() else None)
    extractor.load_locales(workspace_root)
    
    if args.locale not in extractor.locale_data:
        print(f"❌ Error: Locale '{args.locale}' not found")
        return 1
    
    rows = extractor.extract_all(locale=args.locale)
    
    print(f"\n📊 Extracted {len(rows)} copywriting items\n")
    
    # Generate output files
    date_str = datetime.now().strftime('%Y%m%d')
    base_filename = f"{args.prefix}_{date_str}"
    
    if args.format in ['csv', 'both']:
        csv_path = output_dir / f"{base_filename}.csv"
        extractor.generate_csv(csv_path)
    
    if args.format in ['xlsx', 'both']:
        xlsx_path = output_dir / f"{base_filename}.xlsx"
        extractor.generate_xlsx(xlsx_path)
    
    print("\n✨ Done!")
    return 0


if __name__ == "__main__":
    exit(main())
